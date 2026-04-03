"""
src.py  –  Motore di calcolo per il Computo Metrico Estimativo
==============================================================
Versione 2.0 – Strutture dati aggiornate:
  · Tabella Sintesi: progressiva, articolo, descrizione_breve, lotto, WBS, categoria
  · Libretto Misure: commento, simili (n°), lung, larg, alt, quantita_diretta
  · Tipi voce: standard | sovrapprezzo_pct | riferimento
  · Tipi riga misurazione: standard | sottrazione | riferimento_voce

Schema JSON – Voce di Computo
──────────────────────────────
{
  "id": 3,
  "progressiva":      "1.3",
  "articolo":         "SC.02.001",          ← era "codice"
  "descrizione":      "Scavo a sezione...",
  "descrizione_breve":"Scavo sez. obbligata",
  "um":               "mc",
  "prezzo_unitario":  18.50,
  "lotto":            "L1",
  "wbs":              "1.2",
  "categoria":        "Scavi",
  "sottocategoria":   "Meccanici",
  "note":             "",
  "tipo":             "standard",            ← "standard"|"sovrapprezzo_pct"|"riferimento"
  "sovrapprezzo_pct": null,                  ← % se tipo==sovrapprezzo_pct (es. 15.0)
  "rif_voce_id":      null,                  ← id voce padre per sovrapprezzo/riferimento
  "misurazioni": [
    {
      "id":               "m_001",
      "commento":         "Tratto A-B via Roma",   ← era "descrizione"
      "simili":           2.0,                     ← era "parti" (n° pezzi uguali)
      "lung":             15.50,
      "larg":             3.00,
      "alt":              2.50,
      "quantita_diretta": 0.0,                     ← era "quantita" (usata se no dimensioni)
      "tipo_riga":        "standard",              ← "standard"|"sottrazione"|"riferimento_voce"
      "rif_voce_id":      null                     ← per tipo_riga==riferimento_voce
    }
  ],
  "quantita_totale": 232.50,
  "importo":         4301.25
}

Logica Sovrapprezzi (tipo == "sovrapprezzo_pct")
─────────────────────────────────────────────────
  Opzione A – Nella colonna Simili (quantità):
      Inserire il coefficiente percentuale come quantita_diretta (es. 0.15 per 15%).
      Il prezzo unitario della voce viene calcolato come:
          PU_sovrapprezzo = PU_voce_rif * sovrapprezzo_pct / 100
      La quantità è quella dalla voce_rif (copiata tramite riferimento_voce).

  Opzione B – Trasformazione diretta del prezzo (calcolata in calcola_importo):
      importo = importo_voce_rif * sovrapprezzo_pct / 100
      Quantità = quella della voce di riferimento.

Logica Riferimento Voce (tipo_riga == "riferimento_voce")
──────────────────────────────────────────────────────────
  Una riga misurazione con tipo_riga=="riferimento_voce" e rif_voce_id valorizzato
  restituisce la quantita_totale della voce referenziata.
  Supporta catene: A → B → C (con ciclo-detection).
"""

from __future__ import annotations

import io
import json
import re
import uuid
from typing import Any

import pandas as pd

# ══════════════════════════════════════════════════════════════════════════════
# COSTANTI
# ══════════════════════════════════════════════════════════════════════════════

COLONNE_PREZZIARIO = ["CODICE", "DESCRIZIONE", "UM", "PREZZO", "FONTE"]

TIPI_VOCE      = ("standard", "sovrapprezzo_pct", "riferimento")
TIPI_RIGA      = ("standard", "sottrazione", "riferimento_voce")

MISURAZIONE_VUOTA: dict = {
    "id":               "",
    "commento":         "",
    "simili":           1.0,
    "lung":             0.0,
    "larg":             0.0,
    "alt":              0.0,
    "quantita_diretta": 0.0,
    "tipo_riga":        "standard",
    "rif_voce_id":      None,
}


# ══════════════════════════════════════════════════════════════════════════════
# UTILITÀ GENERALI
# ══════════════════════════════════════════════════════════════════════════════

def parse_price(val_str: Any) -> float:
    """Converte stringa prezzo → float. Gestisce '1.234,56' e '1234.56'."""
    if val_str is None:
        return 0.0
    s = str(val_str).strip()
    if re.search(r"\d\.\d{3},\d", s):
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    m = re.search(r"\d+\.?\d*", s)
    return float(m.group()) if m else 0.0


def _safe_str(val: Any) -> str:
    s = str(val or "").strip()
    return "" if s in ("nan", "None", "NaN") else s


def _safe_float(val: Any) -> float:
    try:
        return float(str(val).replace(",", "."))
    except (ValueError, TypeError):
        return 0.0


def _new_mis_id() -> str:
    return f"m_{uuid.uuid4().hex[:8]}"


# ══════════════════════════════════════════════════════════════════════════════
# BACKWARD COMPATIBILITY  – Mapping campi vecchi → nuovi
# ══════════════════════════════════════════════════════════════════════════════

def _normalizza_misurazione(m: dict) -> dict:
    """Migra un dict misurazione vecchio formato → nuovo formato."""
    out = dict(MISURAZIONE_VUOTA)
    out["id"]               = m.get("id") or _new_mis_id()
    out["commento"]         = m.get("commento") or m.get("descrizione", "")
    out["simili"]           = float(m.get("simili") or m.get("parti", 1) or 1)
    out["lung"]             = float(m.get("lung", 0) or 0)
    out["larg"]             = float(m.get("larg", 0) or 0)
    out["alt"]              = float(m.get("alt",  0) or 0)
    out["quantita_diretta"] = float(m.get("quantita_diretta") or m.get("quantita", 0) or 0)
    out["tipo_riga"]        = m.get("tipo_riga", "standard")
    out["rif_voce_id"]      = m.get("rif_voce_id")
    return out


def _normalizza_voce(v: dict) -> dict:
    """Migra una voce vecchio formato → nuovo formato."""
    v.setdefault("progressiva",      "")
    v.setdefault("articolo",         v.get("codice", ""))
    v.setdefault("descrizione_breve","")
    v.setdefault("lotto",            "")
    v.setdefault("tipo",             "standard")
    v.setdefault("sovrapprezzo_pct", None)
    v.setdefault("rif_voce_id",      None)
    # migra misurazioni
    v["misurazioni"] = [_normalizza_misurazione(m) for m in (v.get("misurazioni") or [])]
    return v


# ══════════════════════════════════════════════════════════════════════════════
# LIBRETTO DELLE MISURE  –  Calcolo quantità
# ══════════════════════════════════════════════════════════════════════════════

def quantita_misurazione(
    m: dict,
    computo: list[dict] | None = None,
    _visited: set | None = None,
) -> float:
    """
    Calcola la quantità di una riga misurazione.

    Tipi riga:
      standard          → simili × (lung × larg × alt)  oppure  simili × quantita_diretta
      sottrazione       → valore negativo del calcolo standard
      riferimento_voce  → quantita_totale della voce referenziata (ciclo-safe)
    """
    tipo = m.get("tipo_riga", "standard")

    # ── Riferimento voce ──────────────────────────────────────────────────────
    if tipo == "riferimento_voce" and computo is not None:
        rif_id = m.get("rif_voce_id")
        if rif_id is not None:
            if _visited is None:
                _visited = set()
            if rif_id in _visited:
                return 0.0  # ciclo rilevato
            _visited.add(rif_id)
            rif_voce = next((v for v in computo if v["id"] == rif_id), None)
            if rif_voce is not None:
                return quantita_totale_voce(rif_voce, computo, _visited=_visited)
        return 0.0

    # ── Calcolo dimensionale ──────────────────────────────────────────────────
    simili = float(m.get("simili", 1) or 1)
    lung   = float(m.get("lung",   0) or 0)
    larg   = float(m.get("larg",   0) or 0)
    alt    = float(m.get("alt",    0) or 0)

    if lung or larg or alt:
        l = lung if lung else 1.0
        w = larg if larg else 1.0
        h = alt  if alt  else 1.0
        q = round(simili * l * w * h, 4)
    else:
        q = round(simili * float(m.get("quantita_diretta", 0) or 0), 4)

    return -q if tipo == "sottrazione" else q


def quantita_totale_voce(
    voce: dict,
    computo: list[dict] | None = None,
    _visited: set | None = None,
) -> float:
    """
    Somma le quantità di tutte le righe di misurazione.

    Se tipo == "riferimento": restituisce la quantita_totale della voce padre.
    """
    tipo = voce.get("tipo", "standard")

    # ── Voce di tipo Riferimento (copia quantità da altra voce) ───────────────
    if tipo == "riferimento" and computo is not None:
        rif_id = voce.get("rif_voce_id")
        if rif_id is not None:
            if _visited is None:
                _visited = set()
            vid = voce["id"]
            if vid in _visited:
                return 0.0
            _visited.add(vid)
            padre = next((v for v in computo if v["id"] == rif_id), None)
            if padre is not None:
                return quantita_totale_voce(padre, computo, _visited=_visited)
        return 0.0

    misurazioni = voce.get("misurazioni") or []
    if not misurazioni:
        return float(voce.get("quantita_diretta", 0) or voce.get("quantita_totale", 0) or 0)
    return round(
        sum(quantita_misurazione(m, computo, _visited=set(_visited or ())) for m in misurazioni),
        4,
    )


def calcola_importo(
    voce: dict,
    computo: list[dict] | None = None,
) -> float:
    """
    Calcola l'importo della voce.

    Tipi:
      standard          → quantita_totale × prezzo_unitario
      sovrapprezzo_pct  → importo_voce_rif × sovrapprezzo_pct/100
      riferimento       → quantita_totale (ereditata) × prezzo_unitario
    """
    tipo = voce.get("tipo", "standard")
    qt   = quantita_totale_voce(voce, computo)

    # ── Sovrapprezzo % sull'importo della voce di riferimento ─────────────────
    if tipo == "sovrapprezzo_pct" and computo is not None:
        pct    = float(voce.get("sovrapprezzo_pct", 0) or 0)
        rif_id = voce.get("rif_voce_id")
        if rif_id is not None:
            padre = next((v for v in computo if v["id"] == rif_id), None)
            if padre is not None:
                base = calcola_importo(padre, computo)
                return round(base * pct / 100.0, 2)
        # Fallback: usa quantita_diretta × PU × pct
        pu = float(voce.get("prezzo_unitario", 0) or 0)
        return round(qt * pu * pct / 100.0, 2)

    pu = float(voce.get("prezzo_unitario", 0) or 0)
    return round(qt * pu, 2)


def aggiorna_importi(computo: list[dict]) -> None:
    """Ricalcola quantita_totale e importo per tutte le voci. In-place."""
    # Prima normalizza tutti i formati
    for v in computo:
        _normalizza_voce(v)
    # Poi ricalcola (passando il computo intero per i riferimenti)
    for v in computo:
        v["quantita_totale"] = quantita_totale_voce(v, computo)
        v["importo"]         = calcola_importo(v, computo)


def nuova_misurazione(**kwargs) -> dict:
    m = dict(MISURAZIONE_VUOTA)
    m["id"] = _new_mis_id()
    m.update(kwargs)
    return m


def nuova_voce(next_id: int, **kwargs) -> dict:
    """Crea una voce di computo con struttura completa v2.0."""
    v: dict = {
        "id":               next_id,
        "progressiva":      "",
        "articolo":         "",
        "descrizione":      "",
        "descrizione_breve":"",
        "um":               "",
        "prezzo_unitario":  0.0,
        "lotto":            "",
        "wbs":              "",
        "categoria":        "",
        "sottocategoria":   "",
        "note":             "",
        "tipo":             "standard",
        "sovrapprezzo_pct": None,
        "rif_voce_id":      None,
        "misurazioni":      [nuova_misurazione()],
        "quantita_totale":  0.0,
        "importo":          0.0,
    }
    v.update(kwargs)
    # Mantieni compatibilità: se passato "codice", mettilo in "articolo"
    if "codice" in kwargs and not kwargs.get("articolo"):
        v["articolo"] = kwargs["codice"]
    return v


# ══════════════════════════════════════════════════════════════════════════════
# PROGRESSIVA  –  Assegnazione automatica numeri progressivi
# ══════════════════════════════════════════════════════════════════════════════

def assegna_progressive(computo: list[dict]) -> None:
    """
    Assegna il numero progressivo alle voci in base alla categoria.
    Es: cat "Scavi" → 1.1, 1.2, … | cat "Fondazioni" → 2.1, 2.2, …
    """
    cat_idx:  dict[str, int] = {}
    voc_idx:  dict[str, int] = {}

    for v in computo:
        cat = v.get("categoria") or "—"
        if cat not in cat_idx:
            cat_idx[cat]  = len(cat_idx) + 1
            voc_idx[cat]  = 0
        voc_idx[cat]  += 1
        v["progressiva"] = f"{cat_idx[cat]}.{voc_idx[cat]}"


# ══════════════════════════════════════════════════════════════════════════════
# AGGREGAZIONE E RIEPILOGO WBS
# ══════════════════════════════════════════════════════════════════════════════

def totale_computo(computo: list[dict]) -> float:
    return round(sum(v.get("importo", 0.0) for v in computo), 2)


def riepilogo_wbs(computo: list[dict]) -> pd.DataFrame:
    if not computo:
        return pd.DataFrame(columns=["WBS", "Lotto", "Categoria", "Sottocategoria", "N. voci", "Importo €"])

    rows = [
        {
            "WBS":            v.get("wbs", ""),
            "Lotto":          v.get("lotto", ""),
            "Categoria":      v.get("categoria") or "— Senza categoria —",
            "Sottocategoria": v.get("sottocategoria", ""),
            "Importo":        v.get("importo", 0.0),
        }
        for v in computo
    ]
    df = pd.DataFrame(rows)
    agg = (
        df.groupby(["WBS", "Lotto", "Categoria", "Sottocategoria"])["Importo"]
        .agg(N_voci="count", Importo="sum")
        .reset_index()
        .rename(columns={"Importo": "Importo €", "N_voci": "N. voci"})
        .sort_values("Importo €", ascending=False)
        .reset_index(drop=True)
    )
    return agg


def computo_to_dataframe(computo: list[dict]) -> pd.DataFrame:
    if not computo:
        return pd.DataFrame()
    return pd.DataFrame([
        {
            "Prg.":          v.get("progressiva", ""),
            "Articolo":      v.get("articolo") or v.get("codice", ""),
            "Breve":         str(v.get("descrizione_breve", "") or v.get("descrizione", ""))[:50],
            "Descrizione":   str(v.get("descrizione", "") or "")[:120],
            "UM":            v.get("um", ""),
            "Quantità":      v.get("quantita_totale", 0),
            "P.U. €":        v.get("prezzo_unitario", 0),
            "Importo €":     v.get("importo", 0),
            "Lotto":         v.get("lotto", ""),
            "WBS":           v.get("wbs", ""),
            "Categoria":     v.get("categoria", ""),
            "Tipo":          v.get("tipo", "standard"),
        }
        for v in computo
    ])


# ══════════════════════════════════════════════════════════════════════════════
# PARSING PREZZIARIO DA PDF
# ══════════════════════════════════════════════════════════════════════════════

_RE_CODICE = re.compile(r"^[A-Za-z]{1,3}\.\d{2,3}(?:\.\d{3})?(?:\.[a-z])?$")


def _parse_rows_from_table(table: list, nome: str) -> list[dict]:
    rows = []
    for row in table:
        if row is None or len(row) < 4:
            continue
        code_cell  = str(row[0] or "").strip()
        desc_cell  = str(row[1] or "").strip()
        um_cell    = str(row[2] or "").strip()
        price_cell = str(row[3] or "").strip()

        if code_cell.upper() in ("CODICE", "COD.", "", "-"):
            continue
        if "descrizione" in code_cell.lower():
            continue

        codes  = [c.strip() for c in code_cell.split("\n") if c.strip()]
        prices = [p.strip() for p in price_cell.split("\n") if p.strip()]
        ums    = [u.strip() for u in um_cell.split("\n")    if u.strip()]
        descs  = desc_cell.split("\n")

        for i, cod in enumerate(codes):
            if not _RE_CODICE.match(cod):
                continue
            pr = prices[i] if i < len(prices) else (prices[-1] if prices else "0")
            um = ums[i]    if i < len(ums)    else (ums[-1]    if ums    else "")

            desc_chunk = descs[i].strip() if i < len(descs) else ""
            if len(desc_chunk) < 5 and desc_cell:
                desc_chunk = desc_cell[:300].replace("\n", " ")
            desc_clean = re.sub(r"\s+", " ", desc_chunk).strip()

            price_val = parse_price(pr)
            if price_val <= 0:
                continue

            rows.append({"CODICE": cod, "DESCRIZIONE": desc_clean,
                         "UM": um, "PREZZO": price_val, "FONTE": nome})
    return rows


def _extract_rows_from_text(text: str, nome: str) -> list[dict]:
    rows = []
    pattern = re.compile(
        r"^([A-Za-z]{1,3}\.\d{2,3}(?:\.\d{3})?(?:\.[a-z])?)"
        r"\s+(.+?)\s+([A-Za-z²³/]{1,5})\s+(\d[\d.,]+)",
        re.MULTILINE,
    )
    for m in pattern.finditer(text):
        price_val = parse_price(m.group(4))
        if price_val <= 0:
            continue
        rows.append({
            "CODICE":      m.group(1).strip(),
            "DESCRIZIONE": re.sub(r"\s+", " ", m.group(2)).strip(),
            "UM":          m.group(3).strip(),
            "PREZZO":      price_val,
            "FONTE":       nome,
        })
    return rows


def _finalize_df(rows: list[dict]) -> pd.DataFrame:
    return (
        pd.DataFrame(rows)
        .drop_duplicates(subset=["CODICE"])
        .reset_index(drop=True)
    )


def extract_pdf_prezziario(pdf_bytes: bytes, nome: str) -> pd.DataFrame:
    rows: list[dict] = []
    try:
        import fitz
        with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
            for page in doc:
                try:
                    for tab in page.find_tables():
                        rows.extend(_parse_rows_from_table(tab.extract(), nome))
                except AttributeError:
                    rows.extend(_extract_rows_from_text(page.get_text("text"), nome))
        if rows:
            return _finalize_df(rows)
    except ImportError:
        pass

    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                for table in (page.extract_tables() or []):
                    rows.extend(_parse_rows_from_table(table, nome))
        if rows:
            return _finalize_df(rows)
    except ImportError as exc:
        raise ImportError(
            "Installa almeno uno tra PyMuPDF e pdfplumber:\n"
            "  pip install pymupdf\n  pip install pdfplumber"
        ) from exc

    return pd.DataFrame(columns=COLONNE_PREZZIARIO)


# ══════════════════════════════════════════════════════════════════════════════
# PARSING PREZZIARIO DA XLSX
# ══════════════════════════════════════════════════════════════════════════════

_COL_KEYWORDS: dict[str, list[str]] = {
    "CODICE":      ["TARIFFA", "CODICE", "COD"],
    "DESCRIZIONE": ["DESCRIZIONE", "DESCR"],
    "UM":          ["UM", "U.M.", "UDM", "UNITÀ", "UNITA"],
    "PREZZO":      ["PREZZO", "COSTO"],
}


def _detect_header_row(df: pd.DataFrame) -> int | None:
    keywords = {"tariffa", "codice", "prezzo", "descrizione"}
    for i, row in df.iterrows():
        row_str = " ".join(str(v).lower() for v in row if pd.notna(v))
        if any(kw in row_str for kw in keywords):
            return int(i)
    return None


def _map_columns(columns: list[str]) -> dict[str, str]:
    col_map: dict[str, str] = {}
    for col in columns:
        for standard, kws in _COL_KEYWORDS.items():
            if standard not in col_map and any(k in col.upper() for k in kws):
                col_map[standard] = col
                break
    return col_map


def extract_xlsx_prezziario(xlsx_bytes: bytes, nome: str) -> pd.DataFrame:
    xls  = pd.ExcelFile(io.BytesIO(xlsx_bytes))
    dfs: list[pd.DataFrame] = []

    for sheet in xls.sheet_names:
        raw        = pd.read_excel(io.BytesIO(xlsx_bytes), sheet_name=sheet, header=None)
        header_row = _detect_header_row(raw)
        if header_row is None:
            continue

        raw.columns = raw.iloc[header_row]
        raw         = raw.iloc[header_row + 1:].copy()
        raw.columns = [str(c).strip().upper() for c in raw.columns]
        col_map     = _map_columns(list(raw.columns))

        if not all(k in col_map for k in ("CODICE", "DESCRIZIONE", "PREZZO")):
            continue

        wanted = [k for k in ("CODICE", "DESCRIZIONE", "UM", "PREZZO") if k in col_map]
        sub    = raw[[col_map[k] for k in wanted]].copy()
        sub.columns = wanted
        sub    = sub.dropna(subset=["CODICE", "PREZZO"])
        sub["PREZZO"] = pd.to_numeric(sub["PREZZO"], errors="coerce").fillna(0)
        sub    = sub[sub["PREZZO"] > 0].copy()

        if not sub.empty:
            sub["FONTE"] = nome
            dfs.append(sub)

    if not dfs:
        return pd.DataFrame(columns=COLONNE_PREZZIARIO)
    return (
        pd.concat(dfs, ignore_index=True)
        .drop_duplicates(subset=["CODICE"])
        .reset_index(drop=True)
    )


# ══════════════════════════════════════════════════════════════════════════════
# AGGREGAZIONE PREZZIARI
# ══════════════════════════════════════════════════════════════════════════════

def get_all_voci(prezziari: dict[str, pd.DataFrame]) -> pd.DataFrame:
    dfs = [df for df in prezziari.values() if not df.empty]
    if not dfs:
        return pd.DataFrame(columns=COLONNE_PREZZIARIO)
    return (
        pd.concat(dfs, ignore_index=True)
        .drop_duplicates(subset=["CODICE"])
        .reset_index(drop=True)
    )


def cerca_voce(query: str, df_voci: pd.DataFrame, max_results: int = 30) -> pd.DataFrame:
    if df_voci.empty or not query.strip():
        return df_voci.head(max_results)
    q    = query.strip().upper()
    mask = (
        df_voci["CODICE"].str.upper().str.contains(q, na=False, regex=False)
        | df_voci["DESCRIZIONE"].str.upper().str.contains(q, na=False, regex=False)
    )
    return df_voci[mask].head(max_results)


def lookup_voce_by_codice(codice: str, df_voci: pd.DataFrame) -> dict | None:
    match = df_voci[df_voci["CODICE"].str.upper() == codice.strip().upper()]
    if match.empty:
        return None
    r = match.iloc[0]
    return {
        "codice":      r["CODICE"],
        "descrizione": r["DESCRIZIONE"],
        "um":          r.get("UM", ""),
        "prezzo":      float(r["PREZZO"]),
        "fonte":       r.get("FONTE", ""),
    }


# ══════════════════════════════════════════════════════════════════════════════
# IMPORTAZIONE COMPUTO DA XLSX ESISTENTE
# ══════════════════════════════════════════════════════════════════════════════

def import_computo_from_xlsx(
    xlsx_bytes:   bytes,
    sheet_name:   str,
    header_row:   int,
    col_cat:      int,
    col_sottocat: int,
    col_cod:      int,
    col_desc:     int,
    col_um:       int,
    col_q:        int,
    col_pu:       int,
    start_id:     int = 1,
) -> tuple[list[dict], int]:
    raw        = pd.read_excel(io.BytesIO(xlsx_bytes), sheet_name=sheet_name, header=None)
    dati       = raw.iloc[header_row + 1:].copy()
    voci:      list[dict] = []
    current_id = start_id

    for _, row in dati.iterrows():
        n        = len(row)
        cat      = _safe_str(row.iloc[col_cat])      if col_cat      < n else ""
        sottocat = _safe_str(row.iloc[col_sottocat]) if col_sottocat < n else ""
        cod      = _safe_str(row.iloc[col_cod])      if col_cod      < n else ""
        desc     = _safe_str(row.iloc[col_desc])     if col_desc     < n else ""
        um       = _safe_str(row.iloc[col_um])       if col_um       < n else ""
        q        = _safe_float(row.iloc[col_q])      if col_q        < n else 0.0
        pu       = _safe_float(row.iloc[col_pu])     if col_pu       < n else 0.0

        if not cod and not desc:
            continue

        mis = nuova_misurazione(commento="da XLSX", quantita_diretta=q)
        v   = nuova_voce(
            current_id,
            categoria=cat or "—", sottocategoria=sottocat,
            articolo=cod, descrizione=desc, um=um, prezzo_unitario=pu,
            note="importato da XLSX", misurazioni=[mis],
        )
        v["quantita_totale"] = quantita_totale_voce(v)
        v["importo"]         = calcola_importo(v)
        voci.append(v)
        current_id += 1

    return voci, current_id


# ══════════════════════════════════════════════════════════════════════════════
# SERIALIZZAZIONE JSON
# ══════════════════════════════════════════════════════════════════════════════

def export_json(computo: list[dict], nomi_prezziari: list[str]) -> str:
    return json.dumps(
        {"computo": computo, "prezziari_caricati": nomi_prezziari, "version": "2.0"},
        ensure_ascii=False, indent=2,
    )


def import_json(data: bytes | str) -> dict:
    try:
        obj = json.loads(data)
    except json.JSONDecodeError as exc:
        raise ValueError(f"File JSON non valido: {exc}") from exc
    if "computo" not in obj:
        raise ValueError("Il file JSON non contiene la chiave 'computo'.")
    return {
        "computo":            obj.get("computo", []),
        "prezziari_caricati": obj.get("prezziari_caricati", []),
    }


# ══════════════════════════════════════════════════════════════════════════════
# ESPORTAZIONE EXCEL  (formule vive + riepilogo WBS)
# ══════════════════════════════════════════════════════════════════════════════

def export_excel(computo: list[dict], titolo_progetto: str = "Computo Metrico Estimativo") -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    C_DARK   = "1A1F36"
    C_MED    = "2D3561"
    C_ACCENT = "4A6CF7"
    C_LIGHT  = "E8ECF4"
    C_WHITE  = "FFFFFF"
    C_DGREEN = "1E7E34"
    C_BGROW  = "F5F7FF"

    def _font(bold=False, color=C_DARK, size=10, italic=False):
        return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)
    def _fill(color):
        return PatternFill("solid", fgColor=color)

    thin_s   = Side(style="thin",   color="CCCCCC")
    thick_s  = Side(style="medium", color="999999")
    brd_thin = Border(left=thin_s,  right=thin_s,  top=thin_s,  bottom=thin_s)
    brd_thk  = Border(left=thick_s, right=thick_s, top=thick_s, bottom=thick_s)
    al_c     = Alignment(horizontal="center", vertical="center")
    al_r     = Alignment(horizontal="right",  vertical="center")
    al_l     = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    FMT_EUR = '#,##0.00 "€"'
    FMT_N3  = '#,##0.000'
    FMT_N2  = '#,##0.00'

    # Colonne: A=Prg B=Lotto C=WBS D=Articolo E=Descrizione F=UM
    #          G=Commento H=Simili I=Lung J=Larg K=Alt L=Qtà M=PU N=Importo
    widths = [
        ("A", 6), ("B", 7), ("C", 7), ("D", 12), ("E", 50), ("F", 6),
        ("G", 22), ("H", 7), ("I", 9), ("J", 9), ("K", 9), ("L", 12), ("M", 14), ("N", 15),
    ]
    ws = wb.active
    ws.title        = "Computo Metrico"
    ws.freeze_panes = "A4"
    for col_l, w in widths:
        ws.column_dimensions[col_l].width = w

    # Titolo
    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value     = titolo_progetto.upper()
    c.font      = _font(bold=True, color=C_WHITE, size=14)
    c.fill      = _fill(C_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Intestazioni
    hdrs = ["Prg.", "Lotto", "WBS", "Articolo", "Descrizione / Commento",
            "UM", "Commento mis.", "Simili", "Lung.", "Larg.", "Alt.", "Quantità", "P.U. €", "Importo €"]
    for ci, h in enumerate(hdrs, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font      = _font(bold=True, color=C_WHITE)
        cell.fill      = _fill(C_ACCENT)
        cell.alignment = al_c
        cell.border    = brd_thin
    ws.row_dimensions[2].height = 22

    data_row    = 3
    current_cat = None

    for voce in computo:
        cat  = voce.get("categoria", "") or "—"
        tipo = voce.get("tipo", "standard")

        if cat != current_cat:
            current_cat = cat
            ws.merge_cells(f"A{data_row}:N{data_row}")
            label = f"  {cat.upper()}"
            sc = voce.get("sottocategoria", "")
            if sc:
                label += f"  ›  {sc}"
            cell = ws.cell(row=data_row, column=1, value=label)
            cell.font      = _font(bold=True, color=C_DARK)
            cell.fill      = _fill(C_LIGHT)
            cell.border    = brd_thk
            cell.alignment = Alignment(vertical="center")
            ws.row_dimensions[data_row].height = 18
            data_row += 1

        voce_row    = data_row
        misurazioni = voce.get("misurazioni") or []

        def _vc(col, val=None):
            cell = ws.cell(row=voce_row, column=col, value=val)
            cell.border = brd_thin
            return cell

        _vc(1, voce.get("progressiva", "")).alignment = al_c
        _vc(2, voce.get("lotto", "")).alignment       = al_c
        _vc(3, voce.get("wbs", "")).alignment         = al_c

        c4 = _vc(4, voce.get("articolo") or voce.get("codice", ""))
        c4.font      = _font(color=C_MED)
        c4.alignment = al_c

        # Descrizione + nota sovrapprezzo
        desc = voce.get("descrizione", "")
        if tipo == "sovrapprezzo_pct":
            pct = voce.get("sovrapprezzo_pct", 0)
            desc = f"[SOVR. {pct}%] {desc}"
        elif tipo == "riferimento":
            desc = f"[RIF.→#{voce.get('rif_voce_id','')}] {desc}"

        c5 = _vc(5, desc)
        c5.font      = _font(bold=True)
        c5.alignment = al_l

        _vc(6, voce.get("um", "")).alignment = al_c

        for ci in range(7, 12):
            ws.cell(row=voce_row, column=ci).border = brd_thin

        ws.cell(row=voce_row, column=12).border        = brd_thin
        ws.cell(row=voce_row, column=12).number_format = FMT_N3
        ws.cell(row=voce_row, column=12).alignment     = al_c

        ck = ws.cell(row=voce_row, column=13, value=voce.get("prezzo_unitario", 0))
        ck.font          = _font(bold=True, color=C_MED)
        ck.number_format = FMT_EUR
        ck.border        = brd_thin
        ck.alignment     = al_r

        cl = ws.cell(row=voce_row, column=14)
        cl.font          = _font(bold=True, color=C_DGREEN)
        cl.number_format = FMT_EUR
        cl.border        = brd_thin
        cl.alignment     = al_r
        cl.fill          = _fill("F0FFF4")

        ws.row_dimensions[voce_row].height = 22
        data_row += 1

        mis_start = data_row
        for mis in misurazioni:
            tipo_riga = mis.get("tipo_riga", "standard")

            def _mc(col, val=None):
                cell = ws.cell(row=data_row, column=col, value=val)
                cell.border = brd_thin
                return cell

            for ci in (1, 2, 3, 4, 6):
                _mc(ci)

            commento_txt = mis.get("commento", "")
            if tipo_riga == "sottrazione":
                commento_txt = f"(–) {commento_txt}"
            elif tipo_riga == "riferimento_voce":
                commento_txt = f"[→#{mis.get('rif_voce_id','')}] {commento_txt}"

            cm5 = _mc(7, commento_txt)
            cm5.font      = _font(italic=True, color="666666")
            cm5.alignment = al_l

            _mc(8,  mis.get("simili",           1) or 1).number_format = FMT_N2
            _mc(9,  mis.get("lung",             0) or 0).number_format = FMT_N2
            _mc(10, mis.get("larg",             0) or 0).number_format = FMT_N2
            _mc(11, mis.get("alt",              0) or 0).number_format = FMT_N2
            for ci in (8, 9, 10, 11):
                ws.cell(row=data_row, column=ci).alignment = al_c

            r = data_row
            has_dim = (mis.get("lung") or 0) or (mis.get("larg") or 0) or (mis.get("alt") or 0)
            if has_dim:
                sign = "-" if tipo_riga == "sottrazione" else ""
                fq = (f"={sign}H{r}"
                      f"*IF(I{r}=0,1,I{r})"
                      f"*IF(J{r}=0,1,J{r})"
                      f"*IF(K{r}=0,1,K{r})")
            else:
                q_val = quantita_misurazione(mis)
                fq    = None
                cq    = ws.cell(row=data_row, column=12, value=q_val)
                cq.number_format = FMT_N3
                cq.border        = brd_thin
                cq.alignment     = al_c

            if fq:
                cq = ws.cell(row=data_row, column=12, value=fq)
                cq.number_format = FMT_N3
                cq.border        = brd_thin
                cq.alignment     = al_c

            for ci in (13, 14):
                ws.cell(row=data_row, column=ci).border = brd_thin

            ws.row_dimensions[data_row].height = 17
            data_row += 1

        mis_end = data_row - 1

        j_formula = f"=SUM(L{mis_start}:L{mis_end})" if misurazioni else voce.get("quantita_totale", 0)
        ws.cell(row=voce_row, column=12).value = j_formula
        ws.cell(row=voce_row, column=14).value = f"=L{voce_row}*M{voce_row}"

    # Riga totale
    ws.merge_cells(f"A{data_row}:M{data_row}")
    tl = ws.cell(row=data_row, column=1, value="TOTALE COMPLESSIVO")
    tl.font      = _font(bold=True, color=C_WHITE, size=11)
    tl.fill      = _fill(C_DARK)
    tl.alignment = al_r
    tl.border    = brd_thk

    tv = ws.cell(row=data_row, column=14, value=f"=SUMIF(N3:N{data_row-1},\">0\")")
    tv.font          = _font(bold=True, color=C_WHITE, size=11)
    tv.fill          = _fill(C_DARK)
    tv.number_format = FMT_EUR
    tv.alignment     = al_r
    tv.border        = brd_thk
    ws.row_dimensions[data_row].height = 26

    # Foglio 2 – Riepilogo WBS
    ws2    = wb.create_sheet("Riepilogo WBS")
    df_wbs = riepilogo_wbs(computo)
    rhdrs  = ["WBS", "Lotto", "Categoria", "Sottocategoria", "N. voci", "Importo €"]

    ws2.merge_cells("A1:F1")
    th = ws2["A1"]
    th.value     = f"{titolo_progetto.upper()} – RIEPILOGO WBS"
    th.font      = _font(bold=True, color=C_WHITE, size=12)
    th.fill      = _fill(C_DARK)
    th.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    for cl, w in [("A",8),("B",8),("C",35),("D",30),("E",10),("F",16)]:
        ws2.column_dimensions[cl].width = w

    for ci, h in enumerate(rhdrs, 1):
        cell = ws2.cell(row=2, column=ci, value=h)
        cell.font      = _font(bold=True, color=C_WHITE)
        cell.fill      = _fill(C_ACCENT)
        cell.alignment = al_c
        cell.border    = brd_thin

    for ri, row in df_wbs.iterrows():
        for ci, col_name in enumerate(rhdrs, 1):
            val  = row[col_name]
            cell = ws2.cell(row=ri + 3, column=ci, value=val)
            cell.border = brd_thin
            if col_name == "Importo €":
                cell.number_format = FMT_EUR
                cell.alignment     = al_r
                cell.font          = _font(bold=True, color=C_DGREEN)
            elif col_name == "N. voci":
                cell.alignment = al_c
            else:
                cell.alignment = al_l

    tot_r = len(df_wbs) + 3
    ws2.merge_cells(f"A{tot_r}:E{tot_r}")
    tl2 = ws2.cell(row=tot_r, column=1, value="TOTALE COMPLESSIVO")
    tl2.font = _font(bold=True, color=C_WHITE)
    tl2.fill = _fill(C_DARK)
    tl2.alignment = al_r
    tv2 = ws2.cell(row=tot_r, column=6, value=totale_computo(computo))
    tv2.font = _font(bold=True, color=C_WHITE)
    tv2.fill = _fill(C_DARK)
    tv2.number_format = FMT_EUR
    tv2.alignment = al_r

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# ESPORTAZIONE PDF  (ReportLab)
# ══════════════════════════════════════════════════════════════════════════════

def export_pdf(computo: list[dict], titolo_progetto: str = "Computo Metrico Estimativo") -> bytes:
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import cm, mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError as exc:
        raise ImportError("Installa reportlab: pip install reportlab") from exc

    PAGE_W, PAGE_H = landscape(A4)
    MARGIN = 1.5 * cm

    RL_DARK   = colors.HexColor("#1A1F36")
    RL_ACCENT = colors.HexColor("#4A6CF7")
    RL_LIGHT  = colors.HexColor("#E8ECF4")
    RL_BGROW  = colors.HexColor("#F5F7FF")
    RL_GREEN  = colors.HexColor("#1E7E34")
    RL_GRAY   = colors.HexColor("#888888")
    RL_WHITE  = colors.white

    def _ps(name, font="Helvetica", size=8, color=colors.black, align=TA_LEFT,
            leading=10, bold=False, italic=False):
        fn = font
        if bold and italic: fn += "-BoldOblique"
        elif bold:          fn += "-Bold"
        elif italic:        fn += "-Oblique"
        return ParagraphStyle(name, fontName=fn, fontSize=size,
                              textColor=color, alignment=align, leading=leading)

    s_hdr  = _ps("hdr",  size=7.5, color=RL_WHITE, align=TA_CENTER, bold=True)
    s_desc = _ps("desc", size=8,   color=colors.black, align=TA_LEFT, leading=10)
    s_mit  = _ps("mit",  size=7.5, color=RL_GRAY, align=TA_LEFT, leading=9, italic=True)
    s_num  = _ps("num",  size=8,   color=colors.black, align=TA_RIGHT)
    s_numg = _ps("numg", size=8,   color=RL_GREEN, align=TA_RIGHT, bold=True)
    s_code = _ps("code", font="Courier", size=7.5,
                 color=colors.HexColor("#2D3561"), align=TA_LEFT)
    s_cat  = _ps("cat",  size=9,   color=RL_DARK, align=TA_LEFT, bold=True)
    s_tot  = _ps("tot",  size=9,   color=RL_WHITE, align=TA_RIGHT, bold=True)

    def _p(text, style): return Paragraph(str(text), style)
    def _fmt(val, d=2):
        try:    return f"{float(val):,.{d}f}"
        except: return str(val)

    # Colonne: Prg | Lotto | WBS | Art | Desc/Commento | UM | Sim | Lung | Larg | Alt | Qtà | PU | Imp
    COL_W = [0.6, 0.9, 0.8, 1.8, 8.0, 0.9, 1.1, 1.4, 1.4, 1.4, 1.7, 2.1, 2.4]
    COL_W = [w * cm for w in COL_W]

    hdr_row = [[_p(h, s_hdr) for h in [
        "Prg.", "Lotto", "WBS", "Articolo", "Descrizione / Commento",
        "UM", "Simili", "Lung.", "Larg.", "Alt.", "Quantità", "P.U. €", "Importo €",
    ]]]

    table_data = list(hdr_row)
    ts: list[tuple] = [
        ("BACKGROUND",     (0, 0), (-1, 0),  RL_ACCENT),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [RL_WHITE, RL_BGROW]),
        ("GRID",           (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",     (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING",  (0, 0), (-1, -1), 2),
    ]

    current_cat = None
    abs_r = 1

    for voce in computo:
        cat  = voce.get("categoria", "") or "—"
        tipo = voce.get("tipo", "standard")

        if cat != current_cat:
            current_cat = cat
            label = cat.upper()
            sc = voce.get("sottocategoria", "")
            if sc: label += f"  ›  {sc}"
            table_data.append([_p(f"  {label}", s_cat)] + [""] * 12)
            ts += [
                ("BACKGROUND",    (0, abs_r), (-1, abs_r), RL_LIGHT),
                ("SPAN",          (0, abs_r), (-1, abs_r)),
                ("LINEABOVE",     (0, abs_r), (-1, abs_r), 1, RL_ACCENT),
                ("TOPPADDING",    (0, abs_r), (-1, abs_r), 5),
                ("BOTTOMPADDING", (0, abs_r), (-1, abs_r), 5),
            ]
            abs_r += 1

        qt  = voce.get("quantita_totale", 0)
        pu  = voce.get("prezzo_unitario", 0)
        im  = voce.get("importo", 0)
        desc = voce.get("descrizione", "")
        if tipo == "sovrapprezzo_pct":
            pct  = voce.get("sovrapprezzo_pct", 0)
            desc = f"[SOVR. {pct}%] {desc}"
        elif tipo == "riferimento":
            desc = f"[RIF.→#{voce.get('rif_voce_id','')}] {desc}"

        table_data.append([
            _p(voce.get("progressiva", ""), s_num),
            _p(voce.get("lotto",       ""), s_num),
            _p(voce.get("wbs",         ""), s_num),
            _p(voce.get("articolo") or voce.get("codice", ""), s_code),
            _p(desc,                        s_desc),
            _p(voce.get("um",          ""), s_num),
            "", "", "", "",
            _p(_fmt(qt, 3), s_numg),
            _p(_fmt(pu, 2), s_num),
            _p(_fmt(im, 2), s_numg),
        ])
        ts += [
            ("BACKGROUND", (0, abs_r), (-1, abs_r), colors.HexColor("#F0F3FF")),
            ("LINEBELOW",  (0, abs_r), (-1, abs_r), 0.5, RL_ACCENT),
            ("FONTNAME",   (0, abs_r), (-1, abs_r), "Helvetica-Bold"),
        ]
        abs_r += 1

        for mis in (voce.get("misurazioni") or []):
            tipo_riga = mis.get("tipo_riga", "standard")
            q_mis     = quantita_misurazione(mis)
            commento  = mis.get("commento", "")
            if tipo_riga == "sottrazione":
                commento = f"(–) {commento}"
            elif tipo_riga == "riferimento_voce":
                commento = f"[→#{mis.get('rif_voce_id','')}] {commento}"

            table_data.append([
                "", "", "", "",
                _p(f"    {commento}", s_mit),
                "",
                _p(_fmt(mis.get("simili", 1), 2), s_num),
                _p(_fmt(mis.get("lung",   0), 2), s_num),
                _p(_fmt(mis.get("larg",   0), 2), s_num),
                _p(_fmt(mis.get("alt",    0), 2), s_num),
                _p(_fmt(q_mis,            3), s_num),
                "", "",
            ])
            abs_r += 1

    total = totale_computo(computo)
    table_data.append(
        [_p("TOTALE COMPLESSIVO", s_tot)] + [""] * 11 + [_p(f"€ {total:,.2f}", s_tot)]
    )
    ts += [
        ("BACKGROUND",    (0, abs_r), (-1, abs_r), RL_DARK),
        ("SPAN",          (0, abs_r), (-2, abs_r)),
        ("TOPPADDING",    (0, abs_r), (-1, abs_r), 7),
        ("BOTTOMPADDING", (0, abs_r), (-1, abs_r), 7),
    ]

    main_table = Table(table_data, colWidths=COL_W, repeatRows=1)
    main_table.setStyle(TableStyle(ts))

    buf = io.BytesIO()

    def _on_page(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(RL_GRAY)
        canvas.drawRightString(PAGE_W - MARGIN, 8 * mm, f"Pagina {canvas.getPageNumber()}")
        canvas.drawString(MARGIN, 8 * mm, titolo_progetto)
        canvas.restoreState()

    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=MARGIN, bottomMargin=1.8 * cm,
    )

    title_data  = [[_p(titolo_progetto.upper(),
                        _ps("tt", size=14, color=RL_WHITE, align=TA_CENTER, bold=True))]]
    title_table = Table(title_data, colWidths=[PAGE_W - 2 * MARGIN])
    title_table.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), RL_DARK),
        ("TOPPADDING",    (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
    ]))

    doc.build([title_table, Spacer(1, 0.5 * cm), main_table],
              onFirstPage=_on_page, onLaterPages=_on_page)
    buf.seek(0)
    return buf.getvalue()
